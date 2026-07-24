import html
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from sqlalchemy import text

from app.core.database import Base, engine, SessionLocal
from app.models.koepel import Koepel
from app.models.school import School
from app.models.school_year import Class, SchoolYear
from app.models.school_year import Student
from app.models.user import User
from app.models.goal import Goal
from app.models.observation_goal import ObservationGoal
from app.models.school_goal_domain import SchoolGoalDomain
from app.models.student_observation import StudentObservation
from app.models.theme import Theme
from app.models.activity import Activity, ActivityObservationGoal
from app.models.observation_goal import ObservationGoal
from app.core.security import get_password_hash
from app.core.security import get_password_hash


def reset_database():
    """Drop all tables and recreate them."""
    with engine.connect() as conn:
        conn.execute(text("DROP TABLE IF EXISTS activity_goals CASCADE"))
        conn.commit()
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    print("Database reset complete.")


def seed_koepels():
    """Seed the 3 koepels: OVSG, GO!, and Katholiek Onderwijs Vlaanderen."""
    db = SessionLocal()
    try:
        koepels = [
            {"name": "OVSG", "slug": "ovsg", "is_active": False},
            {"name": "GO!", "slug": "go", "is_active": False},
            {"name": "Katholiek Onderwijs Vlaanderen", "slug": "katholiek-onderwijs-vlaanderen", "is_active": True},
        ]
        
        for koepel_data in koepels:
            existing = db.query(Koepel).filter(Koepel.slug == koepel_data["slug"]).first()
            if not existing:
                koepel = Koepel(
                    name=koepel_data["name"],
                    slug=koepel_data["slug"],
                    is_active=koepel_data["is_active"],
                )
                db.add(koepel)
                print(f"Koepel created: {koepel_data['name']}")
        
        db.commit()
    finally:
        db.close()


def seed_school_and_admin():
    """Create default school and admin user."""
    db = SessionLocal()
    try:
        # Get Katholiek Onderwijs Vlaanderen koepel
        kov = db.query(Koepel).filter(Koepel.slug == "katholiek-onderwijs-vlaanderen").first()
        
        # Create default school
        school = School(
            name="Demo School",
            slug="demo-school",
            is_active=True,
            koepel_id=kov.id if kov else None,
        )
        db.add(school)
        db.commit()
        db.refresh(school)

        # Create default school year 2026-2027
        school_year = SchoolYear(
            school_id=school.id,
            name="2026-2027",
            start_date=date(2025, 9, 1),
            end_date=date(2027, 6, 30),
            is_active=True,
        )
        db.add(school_year)
        db.commit()
        db.refresh(school_year)

        # Create default class 3K
        from app.models.school_year import Class as ClassModel
        from app.models.school_year import Student
 
        class_model = ClassModel(
            school_year_id=school_year.id,
            name="3K",
            class_type="K3",
        )
        db.add(class_model)
        db.commit()
 
        for student_name in [
            "Lena Peeters",
            "Milan Janssens",
            "Noor De Smet",
            "Finn Willems",
            "Amber Claes",
            "Lucas Vermeulen",
            "Sofie Martens",
            "Daan Goossens",
            "Lotte Maes",
            "Thomas Wouters",
            "Eva Peeters",
            "Seppe Janssens",
            "Maud Jacobs",
            "Ruben De Smet",
            "Julie Willems",
        ]:
            db.add(Student(class_id=class_model.id, name=student_name))
        db.commit()
 
        # Create admin user linked to school
        admin = User(
            email="admin@example.com",
            hashed_password=get_password_hash("admin"),
            name="Admin",
            is_superuser=True,
            is_active=True,
            school_id=school.id,
        )
        db.add(admin)
 
        teacher = User(
            email="lieve@example.com",
            hashed_password=get_password_hash("lieve"),
            name="Lieve",
            is_superuser=False,
            is_active=True,
            school_id=school.id,
        )
        db.add(teacher)
        db.commit()

        print(f"School created: {school.name}")
        print(f"School year created: {school_year.name}")
        print(f"Class created: {class_model.name}")
        print("Teacher user created: lieve@example.com / lieve")
        print("Admin user created: admin@example.com / admin")
    finally:
        db.close()



def seed_mow_user():
    """Create demo user (demo@example.com) without koepel selection."""
    db = SessionLocal()
    try:
        # Create demo user (no school yet - will be created after koepel selection)
        mow_user = User(
            email="demo@example.com",
            hashed_password=get_password_hash("demo"),
            name="Demo Uitproberen",
            is_superuser=False,
            is_active=True,
            school_id=None,
            is_pending=False,
            is_demo=True,
        )
        db.add(mow_user)
        db.commit()
        
        print(f"Demo user created: demo@example.com / demo (no koepel selected)")
    finally:
        db.close()

def seed_school_goal_domain_and_goals():
    goals = [
        "Vat een taak of spel spontaan aan",
        "Heeft doorzetting om een taak vol te houden",
        "Werkt nauwkeurig",
        "Kan zich concentreren",
        "Is gemotiveerd en geboeid",
        "Heeft een gezonde exploratiedrang",
        "Werkt rustig",
        "Heeft een goed werktempo",
        "Weet hoe een taak aan te pakken",
        "Kan luisteren zonder afgeleid te zijn",
        "Kan luisteren zonder tussen te komen",
    ]

    db = SessionLocal()
    try:
        school = db.query(School).filter_by(slug="demo-school").first()
        teacher = db.query(User).filter_by(email="lieve@example.com").first() if school else None

        if not school or not teacher:
            print("School goal domain and goals skipped: required school or teacher not found.")
            return

        domain = db.query(SchoolGoalDomain).filter_by(school_id=school.id, name="Betrokkenheid").first()
        if not domain:
            domain = SchoolGoalDomain(school_id=school.id, name="Betrokkenheid")
            db.add(domain)
            db.commit()
            db.refresh(domain)

        for title in goals:
            existing = db.query(ObservationGoal).filter_by(
                school_id=school.id, name=title, domain="Betrokkenheid"
            ).first()
            if not existing:
                db.add(
                    ObservationGoal(
                        school_id=school.id,
                        created_by=teacher.id,
                        name=title,
                        subject=ObservationGoal.SCHOOL_GOALS_SUBJECT,
                        domain="Betrokkenheid",
                        subdomain=None,
                    )
                )

        db.commit()
        print(f"School goal domain seeded: Betrokkenheid ({len(goals)} goals)")
    finally:
        db.close()


def seed_welbevinden_domain_and_goals():
    goals = [
        "Lijkt ontspannen",
        "Vertoont voldoende vitaliteit",
        "Is open en ontvankelijk",
        "Is spontaan",
        "Durft zichzelf te zijn",
        "Heeft zelfvertrouwen",
        "Heeft een positief zelfbeeld",
        "Legt spontaan contact met vriendjes",
        "Legt spontaan contact met andere juffen",
        "Heeft een emotionele stabiliteit (kan plotse en grote veranderinge aan)",
    ]

    db = SessionLocal()
    try:
        school = db.query(School).filter_by(slug="demo-school").first()
        teacher = db.query(User).filter_by(email="lieve@example.com").first() if school else None

        if not school or not teacher:
            print("School goal domain and goals skipped: required school or teacher not found.")
            return

        domain = db.query(SchoolGoalDomain).filter_by(school_id=school.id, name="Welbevinden").first()
        if not domain:
            domain = SchoolGoalDomain(school_id=school.id, name="Welbevinden")
            db.add(domain)
            db.commit()
            db.refresh(domain)

        for title in goals:
            existing = db.query(ObservationGoal).filter_by(
                school_id=school.id, name=title, domain="Welbevinden"
            ).first()
            if not existing:
                db.add(
                    ObservationGoal(
                        school_id=school.id,
                        created_by=teacher.id,
                        name=title,
                        subject=ObservationGoal.SCHOOL_GOALS_SUBJECT,
                        domain="Welbevinden",
                        subdomain=None,
                    )
                )

        db.commit()
        print(f"School goal domain seeded: Welbevinden ({len(goals)} goals)")
    finally:
        db.close()


def _clean_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    # Strip leading apostrophe (from Excel vocabulary fields)
    if text.startswith("'"):
        text = text[1:]
    # Decode HTML entities (e.g. &nbsp; " >)
    text = html.unescape(text)
    # Remove HTML tags
    import re
    text = re.sub(r"<[^>]+>", "", text)
    # Remove zero-width characters
    text = text.replace("\u200b", "").replace("\u200c", "").replace("\u200d", "").replace("\ufeff", "")
    # Normalize special spaces to regular space
    text = text.replace("\u202f", " ").replace("\u2008", " ")
    lines = text.split("\n")
    cleaned_lines = []
    for line in lines:
        line = line.strip()
        if line:
            line = line.lstrip("•").strip()
            if line:
                cleaned_lines.append(line)
    return "\n".join(cleaned_lines) if cleaned_lines else None


def _derive_subject_from_filename(stem: str) -> str:
    name = stem
    for prefix in ["LP-", "Op.stap_", "Op stap_"]:
        if name.startswith(prefix):
            name = name[len(prefix):]
    for suffix in [" (1)", "_com_B_S", "_B_S"]:
        if name.endswith(suffix):
            name = name[: -len(suffix)]

    subject_map = {
        "ned": "Nederlands",
        "wiskunde": "Wiskunde",
        "w_t": "Wetenschap en techniek",
        "aardr": "Aardrijkskunde",
        "frans": "Frans",
        "gesch": "Geschiedenis",
        "ict": "ICT",
        "lele": "Levensleer",
        "lo": "Lichamelijke opvoeding",
        "muvo": "Muziek en visuele opvoeding",
        "rkg": "Religie en levensbeschouwing",
        "v_g": "Vormgeving",
    }
    return _normalize_subject(subject_map.get(name.lower(), name))


def _normalize_subject(value: str) -> str:
    if value.lower() in {"nederlands en communicatie", "nederlands & communicatie", "nederlands-communicatie"}:
        return "Nederlands"
    return value


def seed_vo_goals():
    """Import VO-doelen from Excel."""
    from openpyxl import load_workbook
    from app.models.goal import Goal

    db = SessionLocal()
    try:
        excel_path = Path(__file__).resolve().parent.parent.parent / "AnalysisDev" / "Onderwijsdoelen (3).xlsx"
        wb = load_workbook(excel_path, read_only=True)
        ws = wb["Versie 2.0"]

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[8]:  # Code column
                continue
            code = str(row[8]).strip()
            title = _clean_text(row[9]) or str(row[9]).strip() if row[9] else ""
            description = _clean_text(row[10])
            subject = _normalize_subject(str(row[4]).strip() if row[4] else "")
            level = "K-"
            goal_type = "VO"
            doel_soort = None
            # Column 5 (index 5) = Type doel: "Na te streven" of "Te bereiken"
            target_type_raw = str(row[5]).strip().lower() if row[5] else ""
            if "na te streven" in target_type_raw:
                target_type = "NA_TE_STREVEN"
            elif "te bereiken" in target_type_raw:
                target_type = "TE_BEREIKEN"
            else:
                target_type = None
            parent_goal_id = None
            vo_code = code
            vocabulary = _clean_text(row[12]) if len(row) > 12 else None
            valid_from = row[13] if len(row) > 13 and row[13] else None
            if hasattr(valid_from, "replace"):
                valid_from = valid_from.replace(tzinfo=None)

            existing = db.query(Goal).filter(Goal.code == code).first()
            if existing:
                continue

            goal = Goal(
                code=code,
                title=title,
                description=description,
                subject=subject,
                level=level,
                goal_type=goal_type,
                doel_soort=doel_soort,
                target_type=target_type,
                parent_goal_id=parent_goal_id,
                vo_code=vo_code,
                vocabulary=vocabulary,
                valid_from=valid_from,
            )
            db.add(goal)
            count += 1

        db.commit()
        wb.close()
        print(f"VO-doelen imported: {count}")
    finally:
        db.close()


def _derive_level_from_identifier(identifier: str) -> str:
    """Derive the class level (JK, K2, K3, K-) from an Op Stap identifier."""
    ident = str(identifier)
    # Patterns: .GJK. -> JK, .GK2. or .G K2. -> K2, .GK3. or .G K3. -> K3
    if ".GJK." in ident or ".G JK." in ident:
        return "JK"
    if ".GK2." in ident or ".G K2." in ident:
        return "K2"
    if ".GK3." in ident or ".G K3." in ident:
        return "K3"
    # Patterns: .PF1. -> K-, .PF2. -> K2, .PF3. -> K3
    if ".PF1." in ident:
        return "K-"
    if ".PF2." in ident:
        return "K2"
    if ".PF3." in ident:
        return "K3"
    # Patterns: .GL5. -> K2 (5de leerjaar), .GL6. -> K3 (6de leerjaar)
    if ".GL5." in ident:
        return "K2"
    if ".GL6." in ident:
        return "K3"
    # Patterns: .GL1. -> K2 (1ste leerjaar), .GL2. -> K2, .GL3. -> K3, .GL4. -> K3
    if ".GL1." in ident or ".GL2." in ident:
        return "K2"
    if ".GL3." in ident or ".GL4." in ident:
        return "K3"
    return "K-"


def seed_opstap_goals():
    """Import Op Stap doelen from the generated Excel with minimum goals and examples."""
    from openpyxl import load_workbook

    db = SessionLocal()
    try:
        # First, load all VO goals into a dict keyed by code for linking
        vo_goals = {g.code: g for g in db.query(Goal).filter(Goal.goal_type == "VO").all()}
        print(f"Loaded {len(vo_goals)} VO goals for linking")
        
        # Get Katholiek Onderwijs Vlaanderen koepel for linking
        kov = db.query(Koepel).filter(Koepel.slug == "katholiek-onderwijs-vlaanderen").first()
        print(f"Katholiek Onderwijs Vlaanderen koepel: {kov.name if kov else 'not found'}")

        excel_path = Path(__file__).resolve().parent.parent.parent / "AnalysisDev" / "opstap_naar_minimumdoelen.xlsx"
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active

        # Header: Doel identifier, Doel titel, Pad, Minimumdoel URL, Minimumdoel code, Minimumdoel unieke code, Minimumdoel titel, Voorbeelden
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        print(f"Excel header: {header}")

        count = 0
        linked = 0
        seen_codes = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            goal_identifier = row[0] if len(row) > 0 else None
            if not goal_identifier:
                continue

            code = str(goal_identifier).strip()
            # Handle duplicate codes by appending a counter
            if code in seen_codes:
                counter = 1
                while f"{code}_{counter}" in seen_codes:
                    counter += 1
                code = f"{code}_{counter}"
            seen_codes.add(code)

            title = _clean_text(row[1]) or str(row[1]).strip() if row[1] else ""
            path = str(row[2]).strip() if row[2] else ""
            minimum_goal_code = str(row[4]).strip() if row[4] else ""
            minimum_goal_title = _clean_text(row[6]) or str(row[6]).strip() if row[6] else ""
            voorbeelden = _clean_text(row[7]) if len(row) > 7 else None

            # Derive subject, domain, subdomain, cluster from path
            path_parts = [p.strip() for p in path.split(" > ") if p.strip()]
            subject = _normalize_subject(path_parts[0]) if path_parts else ""
            domain = path_parts[1] if len(path_parts) > 1 else None
            subdomain = path_parts[2] if len(path_parts) > 2 else None
            cluster = path_parts[3] if len(path_parts) > 3 else None

            # Derive level from identifier
            level = _derive_level_from_identifier(goal_identifier)

            # Build description: combine minimum goal title and examples
            description_parts = []
            if minimum_goal_title:
                description_parts.append(minimum_goal_title)
            if voorbeelden:
                description_parts.append(f"Voorbeelden:\n{voorbeelden}")
            description = "\n\n".join(description_parts) if description_parts else None

            # Link to VO goal via minimum_goal_code (remove K- prefix if present)
            vo_code = minimum_goal_code.removeprefix("K-") if minimum_goal_code.startswith("K-") else minimum_goal_code
            if vo_code and vo_code in vo_goals:
                linked += 1

            goal = Goal(
                code=code,
                title=title,
                description=description,
                subject=subject,
                level=level,
                domain=domain,
                subdomain=subdomain,
                cluster=cluster,
                goal_type="OP_STAP",
                minimum_goal_code=minimum_goal_code if minimum_goal_code else None,
                voorbeelden=voorbeelden if voorbeelden else None,
                vo_code=vo_code if vo_code else None,
                koepel_id=kov.id if kov else None,
            )
            db.add(goal)
            count += 1

        db.commit()
        wb.close()
        print(f"Op Stap doelen imported from Excel: {count} (linked to VO: {linked})")
    finally:
        db.close()


def link_demo_observation_goal():
    """Link the demo observation goals to the imported Op Stap goals."""
    demo_goal_specs = [
        # Wiskunde - Getallenkennis
        ("rangtelwoorden", "2.1.GK3.24", "Wiskunde", "Getallenkennis", "Natuurlijke getallen"),
        ("telrij tot 10", "2.1.GK3.20", "Wiskunde", "Getallenkennis", "Natuurlijke getallen"),
        ("aantallen tot 10", "2.1.GK3.15", "Wiskunde", "Getallenkennis", "Natuurlijke getallen"),
        # Wiskunde - Meetkunde
        ("vlakke figuren herkennen", "2.4.GK3.3", "Wiskunde", "Meetkunde", "Vormen"),
        # Nederlands - Lezen
        ("klanken herkennen", "1.2.GK3.3", "Nederlands", "Lezen", "Vlot en vloeiend lezen"),
        ("klank teken koppeling", "1.2.GK3.5", "Nederlands", "Lezen", "Vlot en vloeiend lezen"),
        ("rijmen", "1.2.GK3.1", "Nederlands", "Lezen", "Vlot en vloeiend lezen"),
        ("luisteren naar verhalen", "1.4.GK3.1", "Nederlands", "Mondeling taalgebruik", " Luisterbegrip"),
        # Nederlands - Schrijven
        ("schrijfpatronen", "1.3.GK3.8", "Nederlands", "Schrijven", "Handschrift en digitale vaardigheden"),
        ("binnen lijnen kleuren", "1.3.GK3.9", "Nederlands", "Schrijven", "Handschrift en digitale vaardigheden"),
    ]

    db = SessionLocal()
    try:
        school = db.query(School).filter_by(slug="demo-school").first()
        class_model = db.query(Class).filter_by(name="3K", class_type="K3").first()
        teacher = db.query(User).filter_by(email="lieve@example.com", school_id=school.id if school else None).first()

        if not school or not class_model or not teacher:
            print("Demo observation goals skipped: required school, class or teacher not found.")
            return

        for name, code, subject, domain, subdomain in demo_goal_specs:
            demo_goal = db.query(Goal).filter_by(code=code).first()
            if not demo_goal:
                print(f"Demo observation goal skipped: Op Stap goal {code} not found.")
                continue

            demo_observation_goal = (
                db.query(ObservationGoal)
                .filter(ObservationGoal.school_id == school.id, ObservationGoal.name == name)
                .first()
            )
            if not demo_observation_goal:
                demo_observation_goal = ObservationGoal(
                    school_id=school.id,
                    created_by=teacher.id,
                    name=name,
                    subject=subject,
                    domain=domain,
                    subdomain=subdomain,
                )
                db.add(demo_observation_goal)

            demo_observation_goal.goal_id = demo_goal.id
            print(f"Demo observation goal prepared: {demo_observation_goal.name} -> {demo_goal.code}")

        db.commit()
    finally:
        db.close()


def seed_static_class_observations():
    seeded_class_observations = [
        # Past observations (2025-07-01) - for testing small squares
        ("klanken herkennen", 0, "onvoldoende", date(2025, 7, 1), None),
        ("klanken herkennen", 1, "in_ontwikkeling", date(2025, 7, 1), None),
        ("klank teken koppeling", 2, "voldoende", date(2025, 7, 1), None),
        ("klank teken koppeling", 3, "voorsprong", date(2025, 7, 1), None),
        ("vlakke figuren herkennen", 4, "onvoldoende", date(2025, 7, 1), None),
        ("luisteren naar verhalen", 5, "in_ontwikkeling", date(2025, 7, 1), None),
        ("rangtelwoorden", 6, "voldoende", date(2025, 7, 1), None),
        # Past observations (2025-10-05 to 2025-12-18)
        ("klanken herkennen", 0, "in_ontwikkeling", date(2025, 10, 5), "Voorbeeldcommentaar: herkent de begin- en eindklank."),
        ("klanken herkennen", 1, "voldoende", date(2025, 10, 6), None),
        ("klanken herkennen", 2, "voorsprong", date(2025, 10, 7), None),
        ("klanken herkennen", 3, "in_ontwikkeling", date(2025, 10, 8), None),
        ("klanken herkennen", 4, "voldoende", date(2025, 10, 9), None),
        ("klanken herkennen", 5, "voorsprong", date(2025, 10, 10), None),
        ("klanken herkennen", 6, "onvoldoende", date(2025, 10, 11), None),
        ("klanken herkennen", 7, "in_ontwikkeling", date(2025, 10, 12), None),
        ("klanken herkennen", 8, "voldoende", date(2025, 10, 13), None),
        ("klanken herkennen", 9, "voorsprong", date(2025, 10, 14), None),
        ("klanken herkennen", 10, "in_ontwikkeling", date(2025, 10, 15), None),
        ("klanken herkennen", 11, "voldoende", date(2025, 10, 16), None),
        ("klanken herkennen", 12, "voorsprong", date(2025, 10, 17), None),
        ("klanken herkennen", 13, "onvoldoende", date(2025, 10, 18), None),
        ("klanken herkennen", 14, "in_ontwikkeling", date(2025, 10, 19), None),
        ("klank teken koppeling", 0, "voldoende", date(2025, 10, 20), None),
        ("klank teken koppeling", 1, "voorsprong", date(2025, 10, 21), None),
        ("klank teken koppeling", 2, "in_ontwikkeling", date(2025, 10, 22), None),
        ("klank teken koppeling", 3, "voldoende", date(2025, 10, 23), None),
        ("klank teken koppeling", 4, "voorsprong", date(2025, 10, 24), None),
        ("klank teken koppeling", 5, "onvoldoende", date(2025, 10, 25), None),
        ("klank teken koppeling", 6, "in_ontwikkeling", date(2025, 10, 26), None),
        ("klank teken koppeling", 7, "voldoende", date(2025, 10, 27), None),
        ("klank teken koppeling", 8, "voorsprong", date(2025, 10, 28), None),
        ("klank teken koppeling", 9, "in_ontwikkeling", date(2025, 10, 29), None),
        ("klank teken koppeling", 10, "voldoende", date(2025, 10, 30), None),
        ("klank teken koppeling", 11, "voorsprong", date(2025, 10, 31), None),
        ("klank teken koppeling", 12, "onvoldoende", date(2025, 11, 1), None),
        ("klank teken koppeling", 13, "in_ontwikkeling", date(2025, 11, 2), None),
        ("klank teken koppeling", 14, "voldoende", date(2025, 11, 3), None),
        ("luisteren naar verhalen", 0, "voorsprong", date(2025, 11, 4), None),
        ("luisteren naar verhalen", 1, "in_ontwikkeling", date(2025, 11, 5), None),
        ("luisteren naar verhalen", 2, "voldoende", date(2025, 11, 6), None),
        ("luisteren naar verhalen", 3, "voorsprong", date(2025, 11, 7), None),
        ("luisteren naar verhalen", 4, "onvoldoende", date(2025, 11, 8), None),
        ("luisteren naar verhalen", 5, "in_ontwikkeling", date(2025, 11, 9), None),
        ("luisteren naar verhalen", 6, "voldoende", date(2025, 11, 10), None),
        ("luisteren naar verhalen", 7, "voorsprong", date(2025, 11, 11), None),
        ("luisteren naar verhalen", 8, "in_ontwikkeling", date(2025, 11, 12), None),
        ("luisteren naar verhalen", 9, "voldoende", date(2025, 11, 13), None),
        ("luisteren naar verhalen", 10, "voorsprong", date(2025, 11, 14), None),
        ("luisteren naar verhalen", 11, "onvoldoende", date(2025, 11, 15), None),
        ("luisteren naar verhalen", 12, "in_ontwikkeling", date(2025, 11, 16), None),
        ("luisteren naar verhalen", 13, "voldoende", date(2025, 11, 17), None),
        ("luisteren naar verhalen", 14, "voorsprong", date(2025, 11, 18), None),
        ("rangtelwoorden", 0, "voldoende", date(2025, 11, 19), None),
        ("rangtelwoorden", 1, "voorsprong", date(2025, 11, 20), None),
        ("rangtelwoorden", 2, "in_ontwikkeling", date(2025, 11, 21), None),
        ("rangtelwoorden", 3, "voldoende", date(2025, 11, 22), None),
        ("rangtelwoorden", 4, "voorsprong", date(2025, 11, 23), None),
        ("rangtelwoorden", 5, "onvoldoende", date(2025, 11, 24), None),
        ("rangtelwoorden", 6, "in_ontwikkeling", date(2025, 11, 25), None),
        ("rangtelwoorden", 7, "voldoende", date(2025, 11, 26), None),
        ("rangtelwoorden", 8, "voorsprong", date(2025, 11, 27), None),
        ("rangtelwoorden", 9, "in_ontwikkeling", date(2025, 11, 28), None),
        ("rangtelwoorden", 10, "voldoende", date(2025, 11, 29), None),
        ("rangtelwoorden", 11, "voorsprong", date(2025, 11, 30), None),
        ("rangtelwoorden", 12, "onvoldoende", date(2025, 12, 1), None),
        ("rangtelwoorden", 13, "in_ontwikkeling", date(2025, 12, 2), None),
        ("rangtelwoorden", 14, "voldoende", date(2025, 12, 3), None),
        ("vlakke figuren herkennen", 0, "voorsprong", date(2025, 12, 4), None),
        ("vlakke figuren herkennen", 1, "in_ontwikkeling", date(2025, 12, 5), None),
        ("vlakke figuren herkennen", 2, "voldoende", date(2025, 12, 6), None),
        ("vlakke figuren herkennen", 3, "voorsprong", date(2025, 12, 7), None),
        ("vlakke figuren herkennen", 4, "onvoldoende", date(2025, 12, 8), None),
        ("vlakke figuren herkennen", 5, "in_ontwikkeling", date(2025, 12, 9), None),
        ("vlakke figuren herkennen", 6, "voldoende", date(2025, 12, 10), None),
        ("vlakke figuren herkennen", 7, "voorsprong", date(2025, 12, 11), None),
        ("vlakke figuren herkennen", 8, "in_ontwikkeling", date(2025, 12, 12), None),
        ("vlakke figuren herkennen", 9, "voldoende", date(2025, 12, 13), None),
        ("vlakke figuren herkennen", 10, "voorsprong", date(2025, 12, 14), None),
        ("vlakke figuren herkennen", 11, "onvoldoende", date(2025, 12, 15), None),
        ("vlakke figuren herkennen", 12, "in_ontwikkeling", date(2025, 12, 16), None),
        ("vlakke figuren herkennen", 13, "voldoende", date(2025, 12, 17), None),
        ("vlakke figuren herkennen", 14, "voorsprong", date(2025, 12, 18), None),
    ]

    db = SessionLocal()
    try:
        school = db.query(School).filter_by(slug="demo-school").first()
        class_model = db.query(Class).filter_by(name="3K", class_type="K3").first()
        teacher = db.query(User).filter_by(email="lieve@example.com", school_id=school.id if school else None).first()
        students = (
            db.query(Student)
            .filter(Student.class_id == class_model.id)
            .order_by(Student.name)
            .all()
            if class_model
            else []
        )

        if not school or not class_model or not teacher or not students:
            print("Static class observations skipped: required school, class, teacher or students not found.")
            return

        observation_goals = {
            goal.name: goal
            for goal in db.query(ObservationGoal)
            .filter(ObservationGoal.school_id == school.id)
            .all()
        }

        for goal_name, student_index, status, observation_date, comment in seeded_class_observations:
            if student_index >= len(students) or goal_name not in observation_goals:
                continue

            observation_goal = observation_goals[goal_name]
            student = students[student_index]
            existing = (
                db.query(StudentObservation)
                .filter(
                    StudentObservation.observation_goal_id == observation_goal.id,
                    StudentObservation.student_id == student.id,
                )
                .first()
            )
            if existing:
                continue

            db.add(
                StudentObservation(
                    school_id=school.id,
                    observation_goal_id=observation_goal.id,
                    student_id=student.id,
                    observed_by=teacher.id,
                    status=status,
                    observation_date=observation_date,
                    comment=comment,
                )
            )

        db.commit()
        print(f"Static class observations seeded: {len(seeded_class_observations)} observations.")
    finally:
        db.close()


def seed_themes():
    db = SessionLocal()
    try:
        defaults = [
            ("De appel", "Een thema rond het fruit de appel, gebruikt voor verschillende leeractiviteiten."),
            ("De mol", "Een speels thema waarin de mol centraal staat voor avontuurlijke opdrachten."),
        ]
        for name, description in defaults:
            existing = db.query(Theme).filter(Theme.name == name).first()
            if not existing:
                db.add(Theme(name=name, description=description))
        db.commit()
        print("Default themes seeded.")
    finally:
        db.close()


def seed_activities(school_id: int, teacher_id: int):
    db = SessionLocal()
    try:
        theme = db.query(Theme).filter(Theme.name == "De appel").first()
        if not theme:
            return
        school = db.query(School).filter(School.id == school_id).first()
        if not school:
            return

        opstap_goals = db.query(Goal).filter(Goal.goal_type == "OP_STAP", Goal.subject == "Wiskunde").limit(2).all()
        if not opstap_goals:
            return

        activity = Activity(
            school_id=school_id,
            name="Rekenspelletjes met appels",
            description="Leuke activiteiten met appels om rekenvaardigheid te oefenen.",
            theme_id=theme.id,
        )
        db.add(activity)
        db.commit()
        db.refresh(activity)

        for goal in opstap_goals:
            observation_goal = ObservationGoal(
                school_id=school_id,
                created_by=teacher_id,
                name=goal.title,
                subject=goal.subject,
                domain=goal.domain,
                subdomain=goal.subdomain,
                goal_id=goal.id,
            )
            db.add(observation_goal)
            db.commit()
            db.refresh(observation_goal)

            link = ActivityObservationGoal(
                activity_id=activity.id,
                observation_goal_id=observation_goal.id,
                label=goal.title,
                observe=False,
            )
            db.add(link)

        db.commit()
        print("Default activities seeded.")
    finally:
        db.close()


if __name__ == "__main__":
    reset_database()
    seed_koepels()
    seed_school_and_admin()
    seed_mow_user()
    seed_school_goal_domain_and_goals()
    seed_welbevinden_domain_and_goals()
    seed_vo_goals()
    seed_opstap_goals()
    seed_themes()
    seed_activities(school_id=1, teacher_id=2)
    link_demo_observation_goal()
    seed_static_class_observations()
