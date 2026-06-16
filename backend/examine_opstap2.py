from pathlib import Path
from openpyxl import load_workbook

base = Path("/home/davidlinux/PycharmProjects/ObsApp/AnalysisDev/op_stap")

for fname in ["LP-Ned_com_B_S.xlsx", "LP-wiskunde_B_S (1).xlsx"]:
    path = base / fname
    print(f"\n=== {fname} ===")
    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    
    # Count LfMD values
    lfmd_counts = {}
    doel_soort_counts = {}
    md_samples = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[4]:  # Code column
            continue
        lfmd = str(row[1]).strip() if row[1] else ""
        doel_soort = str(row[0]).strip() if row[0] else ""
        md = str(row[3]).strip() if row[3] else ""
        lfmd_counts[lfmd] = lfmd_counts.get(lfmd, 0) + 1
        doel_soort_counts[doel_soort] = doel_soort_counts.get(doel_soort, 0) + 1
        if lfmd not in md_samples:
            md_samples[lfmd] = md
    
    print(f"LfMD distribution: {lfmd_counts}")
    print(f"Doel-soort distribution: {doel_soort_counts}")
    print(f"Sample MD per LfMD: {md_samples}")
    
    # Show some K- rows with their MD and Code
    print("Sample K- rows (first 5):")
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[4]:
            continue
        lfmd = str(row[1]).strip() if row[1] else ""
        if lfmd == "K-":
            print(f"  MD={row[3]}, Code={row[4]}, Doelsoort={row[0]}, Titel={str(row[9])[:60] if row[9] else ''}")
            count += 1
            if count >= 5:
                break
    wb.close()
