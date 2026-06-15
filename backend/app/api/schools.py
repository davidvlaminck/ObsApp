from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, status, UploadFile
from fastapi.responses import StreamingResponse

from app.api.auth import get_current_school_user, get_current_superuser
from app.core.database import get_db
from app.repositories.school_repository import SchoolRepository
from app.repositories.school_year_repository import ClassRepository, SchoolYearRepository, StudentRepository
from app.schemas.school import (
    ClassCreate,
    ClassResponse,
    SchoolCreate,
    SchoolResponse,
    SchoolYearCreate,
    SchoolYearResponse,
    StudentBulkUploadResult,
    StudentConfirmImport,
    StudentConfirmItem,
    StudentPreviewItem,
    StudentPreviewResult,
    StudentResponse,
    TeacherClassLink,
)
from app.schemas.user import UserResponse

router = APIRouter(prefix="/schools", tags=["schools"])


@router.post("", response_model=SchoolResponse, status_code=status.HTTP_201_CREATED)
@router.post("/", response_model=SchoolResponse, status_code=status.HTTP_201_CREATED)
def create_school(
    payload: SchoolCreate,
    db=Depends(get_db),
    current_superuser: UserResponse = Depends(get_current_superuser),
):
    school_repo = SchoolRepository(db)
    if payload.slug and school_repo.get_by_slug(payload.slug):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Deze slug bestaat al")
    return school_repo.to_response(school_repo.create(payload.name, payload.slug, payload.is_active))


@router.get("", response_model=list[SchoolResponse])
@router.get("/", response_model=list[SchoolResponse])
def list_schools(
    db=Depends(get_db),
    current_superuser: UserResponse = Depends(get_current_superuser),
):
    school_repo = SchoolRepository(db)
    return [school_repo.to_response(school) for school in school_repo.get_all()]


@router.get("/{school_id}/school-years", response_model=list[SchoolYearResponse])
def list_school_years(
    school_id: int,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    if current_user.school_id != school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")
    school_repo = SchoolRepository(db)
    if not school_repo.get_by_id(school_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="School niet gevonden")
    year_repo = SchoolYearRepository(db)
    return [year_repo.to_response(year) for year in year_repo.get_by_school(school_id)]


@router.post("/{school_id}/school-years", response_model=SchoolYearResponse, status_code=status.HTTP_201_CREATED)
def create_school_year(
    school_id: int,
    payload: SchoolYearCreate,
    db=Depends(get_db),
    current_superuser: UserResponse = Depends(get_current_superuser),
):
    school_repo = SchoolRepository(db)
    if not school_repo.get_by_id(school_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="School niet gevonden")
    year_repo = SchoolYearRepository(db)
    school_year = year_repo.create(school_id, payload.name, payload.start_date, payload.end_date, payload.is_active)
    return year_repo.to_response(school_year)


@router.post("/school-years/{school_year_id}/activate", response_model=SchoolYearResponse)
def activate_school_year(
    school_year_id: int,
    db=Depends(get_db),
    current_superuser: UserResponse = Depends(get_current_superuser),
):
    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(school_year_id)
    if not school_year:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Schooljaar niet gevonden")
    return year_repo.to_response(year_repo.set_active(school_year_id))


@router.get("/school-years/{school_year_id}/classes", response_model=list[ClassResponse])
def list_classes(
    school_year_id: int,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(school_year_id)
    if not school_year:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Schooljaar niet gevonden")
    if current_user.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")
    class_repo = ClassRepository(db)
    return [class_repo.class_to_response(cls) for cls in class_repo.get_classes_by_school_year(school_year_id)]


@router.post("/school-years/{school_year_id}/classes", response_model=ClassResponse, status_code=status.HTTP_201_CREATED)
def create_class(
    school_year_id: int,
    payload: ClassCreate,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(school_year_id)
    if not school_year:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Schooljaar niet gevonden")
    if current_user.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")
    class_repo = ClassRepository(db)
    class_model = class_repo.create_class(school_year_id, payload.name, payload.class_type)
    return class_repo.class_to_response(class_model)


@router.get("/classes/{class_id}/students", response_model=list[StudentResponse])
def list_students(
    class_id: int,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    class_repo = ClassRepository(db)
    student_repo = StudentRepository(db)
    class_model = class_repo.get_class_by_id(class_id)
    if not class_model:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Klas niet gevonden")
    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(class_model.school_year_id)
    if current_user.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")
    students = student_repo.get_by_class(class_id)
    return [student_repo.to_response(s) for s in students]


@router.get("/school-years/{school_year_id}/classes/template")
def download_student_template(
    school_year_id: int,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    """Download Excel template for bulk student upload."""
    from openpyxl import Workbook

    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(school_year_id)
    if not school_year:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Schooljaar niet gevonden")
    if current_user.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")

    wb = Workbook()
    ws = wb.active
    ws.title = "Leerlingen"

    # Headers
    ws["A1"] = "Klas"
    ws["B1"] = "Leerling"

    # Example row
    ws["A2"] = "1K"
    ws["B2"] = "Pieter Janssens"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leerlingen_template.xlsx"},
    )


@router.post("/school-years/{school_year_id}/students/bulk", response_model=StudentBulkUploadResult)
def upload_students(
    school_year_id: int,
    file: UploadFile = File(...),
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    """Upload Excel file to bulk add students to classes."""
    from openpyxl import load_workbook

    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(school_year_id)
    if not school_year:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Schooljaar niet gevonden")
    if current_user.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")

    class_repo = ClassRepository(db)
    student_repo = StudentRepository(db)

    # Get all classes for this school year
    classes = class_repo.get_classes_by_school_year(school_year_id)
    class_names = {cls.name: cls for cls in classes}

    # Load the Excel file
    wb = load_workbook(file.file)
    ws = wb.active

    errors = []
    created_count = 0
    total_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0] or not row[1]:
            continue

        class_name = str(row[0]).strip()
        student_name = str(row[1]).strip()
        total_count += 1

        if class_name not in class_names:
            errors.append(f"Rij {row_idx}: Klas '{class_name}' bestaat niet")
            continue

        if not student_name:
            errors.append(f"Rij {row_idx}: Leerlingnaam is leeg")
            continue

        try:
            student_repo.create(class_names[class_name].id, student_name)
            created_count += 1
        except Exception as e:
            errors.append(f"Rij {row_idx}: Fout bij aanmaken leerling: {str(e)}")

    return StudentBulkUploadResult(total=total_count, created=created_count, errors=errors)


@router.post("/school-years/{school_year_id}/students/preview", response_model=StudentPreviewResult)
def preview_students(
    school_year_id: int,
    file: UploadFile = File(...),
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    """Parse Excel file and return preview without saving."""
    from openpyxl import load_workbook

    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(school_year_id)
    if not school_year:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Schooljaar niet gevonden")
    if current_user.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")

    class_repo = ClassRepository(db)
    student_repo = StudentRepository(db)

    # Get all classes for this school year
    classes = class_repo.get_classes_by_school_year(school_year_id)
    class_names = {cls.name: cls for cls in classes}

    # Get existing students per class for duplicate checking
    existing_students: dict[int, set[str]] = {}
    for cls in classes:
        students = student_repo.get_by_class(cls.id)
        existing_students[cls.id] = {s.name.lower() for s in students}

    # Load the Excel file
    wb = load_workbook(file.file)
    ws = wb.active

    items = []
    valid_count = 0
    invalid_count = 0
    # Track duplicates within the import file itself
    seen_in_import: dict[int, set[str]] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0] or not row[1]:
            continue

        class_name = str(row[0]).strip()
        student_name = str(row[1]).strip()

        is_valid = True
        error = None

        if class_name not in class_names:
            is_valid = False
            error = f"Klas '{class_name}' bestaat niet"
        elif not student_name:
            is_valid = False
            error = "Leerlingnaam is leeg"
        else:
            class_id = class_names[class_name].id
            # Check for duplicate within the import file
            if class_id not in seen_in_import:
                seen_in_import[class_id] = set()
            if student_name.lower() in seen_in_import[class_id]:
                is_valid = False
                error = f"Leerling '{student_name}' is dubbel in dit bestand"
            # Check for duplicate in existing database
            elif student_name.lower() in existing_students.get(class_id, set()):
                is_valid = False
                error = f"Leerling '{student_name}' bestaat al in deze klas"
            else:
                seen_in_import[class_id].add(student_name.lower())

        if is_valid:
            valid_count += 1
        else:
            invalid_count += 1

        items.append(StudentPreviewItem(
            class_name=class_name,
            student_name=student_name,
            row_number=row_idx,
            is_valid=is_valid,
            error=error,
        ))

    return StudentPreviewResult(items=items, valid_count=valid_count, invalid_count=invalid_count)


@router.post("/school-years/{school_year_id}/students/confirm", response_model=StudentBulkUploadResult)
def confirm_students(
    school_year_id: int,
    payload: StudentConfirmImport,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    """Confirm and import validated students."""
    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(school_year_id)
    if not school_year:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Schooljaar niet gevonden")
    if current_user.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")

    class_repo = ClassRepository(db)
    student_repo = StudentRepository(db)

    # Get all classes for this school year
    classes = class_repo.get_classes_by_school_year(school_year_id)
    class_names = {cls.name: cls for cls in classes}

    errors = []
    created_count = 0
    total_count = len(payload.items)

    # Track duplicates within the import items
    seen_in_import: dict[int, set[str]] = {}

    for item in payload.items:
        class_name = item.class_name.strip()
        student_name = item.student_name.strip()

        if class_name not in class_names:
            errors.append(f"Klas '{class_name}': Klas bestaat niet")
            continue

        if not student_name:
            errors.append(f"Leerling '{student_name}': Naam is leeg")
            continue

        class_id = class_names[class_name].id

        # Check for duplicate within the import items
        if class_id not in seen_in_import:
            seen_in_import[class_id] = set()
        if student_name.lower() in seen_in_import[class_id]:
            errors.append(f"Leerling '{student_name}' is dubbel in dit bestand")
            continue

        # Check for duplicate in existing database
        existing = student_repo.get_by_class(class_id)
        if any(s.name.lower() == student_name.lower() for s in existing):
            errors.append(f"Leerling '{student_name}' bestaat al in deze klas")
            continue

        seen_in_import[class_id].add(student_name.lower())

        try:
            student_repo.create(class_id, student_name)
            created_count += 1
        except Exception as e:
            errors.append(f"Fout bij aanmaken leerling '{student_name}': {str(e)}")

    return StudentBulkUploadResult(total=total_count, created=created_count, errors=errors)


@router.post("/students/{student_id}/image", response_model=StudentResponse)
def upload_student_image(
    student_id: int,
    file: UploadFile = File(...),
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    """Upload an image for a student."""
    import os
    import uuid

    student_repo = StudentRepository(db)
    student = student_repo.get_by_id(student_id)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Leerling niet gevonden")

    # Get class to check permissions
    class_repo = ClassRepository(db)
    class_model = class_repo.get_class_by_id(student.class_id)
    if not class_model:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Klas niet gevonden")

    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(class_model.school_year_id)
    if current_user.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")

    # Create upload directory if it doesn't exist
    upload_dir = "uploads/students"
    os.makedirs(upload_dir, exist_ok=True)

    # Generate unique filename
    file_ext = file.filename.split('.')[-1] if file.filename else 'jpg'
    filename = f"{uuid.uuid4()}.{file_ext}"
    file_path = os.path.join(upload_dir, filename)

    # Save file
    with open(file_path, "wb") as f:
        f.write(file.file.read())

    # Update student
    student_repo.update_image(student_id, f"/uploads/students/{filename}")
    return student_repo.to_response(student_repo.get_by_id(student_id))


@router.delete("/students/{student_id}", response_model=StudentResponse)
def delete_student(
    student_id: int,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    """Delete a student."""
    student_repo = StudentRepository(db)
    student = student_repo.get_by_id(student_id)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Leerling niet gevonden")

    # Get class to check permissions
    class_repo = ClassRepository(db)
    class_model = class_repo.get_class_by_id(student.class_id)
    if not class_model:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Klas niet gevonden")

    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(class_model.school_year_id)
    if current_user.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")

    # Delete image file if exists
    if student.image_path:
        import os
        full_path = student.image_path
        if full_path.startswith('/'):
            full_path = full_path[1:]
        if os.path.exists(full_path):
            os.remove(full_path)

    student_repo.delete(student_id)
    return StudentResponse(
        id=student.id,
        class_id=student.class_id,
        name=student.name,
        image_path=student.image_path,
        created_at=student.created_at,
    )


@router.get("/classes/{class_id}/teachers", response_model=list[UserResponse])
def list_class_teachers(
    class_id: int,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    class_repo = ClassRepository(db)
    class_model = class_repo.get_class_by_id(class_id)
    if not class_model:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Klas niet gevonden")
    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(class_model.school_year_id)
    if current_user.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")
    return [UserResponse(id=u.id, email=u.email, name=u.name, is_active=u.is_active, is_superuser=u.is_superuser, is_pending=u.is_pending, school_id=u.school_id) for u in class_model.teachers]


@router.post("/classes/{class_id}/teachers", response_model=list[UserResponse], status_code=status.HTTP_201_CREATED)
def add_teacher_to_class(
    class_id: int,
    payload: TeacherClassLink,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    class_repo = ClassRepository(db)
    class_model = class_repo.get_class_by_id(class_id)
    if not class_model:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Klas niet gevonden")
    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(class_model.school_year_id)
    if current_user.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")
    from app.models.user import User as UserModel
    teacher = db.query(UserModel).filter(UserModel.id == payload.teacher_id).first()
    if not teacher:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Leerkracht niet gevonden")
    if teacher.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Leerkracht behoort niet tot deze school")
    class_repo.add_teacher_to_class(class_id, payload.teacher_id)
    return [UserResponse(id=u.id, email=u.email, name=u.name, is_active=u.is_active, is_superuser=u.is_superuser, is_pending=u.is_pending, school_id=u.school_id) for u in class_repo.get_class_by_id(class_id).teachers]


@router.delete("/classes/{class_id}/teachers/{teacher_id}", response_model=list[UserResponse])
def remove_teacher_from_class(
    class_id: int,
    teacher_id: int,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_school_user),
):
    class_repo = ClassRepository(db)
    class_model = class_repo.get_class_by_id(class_id)
    if not class_model:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Klas niet gevonden")
    year_repo = SchoolYearRepository(db)
    school_year = year_repo.get_by_id(class_model.school_year_id)
    if current_user.school_id != school_year.school_id and not current_user.is_superuser:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Niet bevoegd")
    class_repo.remove_teacher_from_class(class_id, teacher_id)
    return [UserResponse(id=u.id, email=u.email, name=u.name, is_active=u.is_active, is_superuser=u.is_superuser, is_pending=u.is_pending, school_id=u.school_id) for u in class_repo.get_class_by_id(class_id).teachers]
