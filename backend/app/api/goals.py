from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, status
from openpyxl import load_workbook

from app.api.auth import get_current_superuser
from app.core.database import get_db
from app.repositories.goal_repository import GoalRepository
from app.schemas.goal import GoalCreate, GoalResponse, GoalUpdate
from app.schemas.user import UserResponse


router = APIRouter(prefix="/goals", tags=["goals"])


def _clean_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    # Clean up bullet points and extra whitespace
    lines = text.split("\n")
    cleaned_lines = []
    for line in lines:
        line = line.strip()
        if line:
            # Remove leading bullet characters and whitespace
            line = line.lstrip("•").strip()
            if line:
                cleaned_lines.append(line)
    return "\n".join(cleaned_lines) if cleaned_lines else None


def _parse_vo_excel(file_path: str) -> list[dict]:
    wb = load_workbook(file_path, read_only=True)
    ws = wb["Versie 2.0"]
    goals = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[8]:  # Code column
            continue
        code = str(row[8]).strip()
        title = str(row[9]).strip() if row[9] else ""
        description = _clean_text(row[10])
        subject = str(row[4]).strip() if row[4] else ""
        level = "K-"  # VO kleuteronderwijs
        goal_type = "VO"
        doel_soort = None
        parent_goal_id = None
        vo_code = code
        vocabulary = _clean_text(row[12]) if len(row) > 12 else None
        valid_from = row[13] if len(row) > 13 and row[13] else None
        if isinstance(valid_from, datetime):
            valid_from = valid_from.replace(tzinfo=None)
        goals.append({
            "code": code,
            "title": title,
            "description": description,
            "subject": subject,
            "level": level,
            "goal_type": goal_type,
            "doel_soort": doel_soort,
            "parent_goal_id": parent_goal_id,
            "vo_code": vo_code,
            "vocabulary": vocabulary,
            "valid_from": valid_from,
        })
    wb.close()
    return goals


@router.get("", response_model=list[GoalResponse])
def list_goals(
    subject: str | None = None,
    goal_type: str | None = None,
    level: str | None = None,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_superuser),
):
    repo = GoalRepository(db)
    goals = repo.get_all(subject=subject, goal_type=goal_type, level=level)
    return [repo.to_response(g) for g in goals]


@router.get("/subjects", response_model=list[str])
def list_subjects(
    goal_type: str | None = None,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_superuser),
):
    repo = GoalRepository(db)
    return repo.get_subjects(goal_type=goal_type)


@router.get("/{goal_id}", response_model=GoalResponse)
def get_goal(
    goal_id: int,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_superuser),
):
    repo = GoalRepository(db)
    goal = repo.get_by_id(goal_id)
    if not goal:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Doel niet gevonden")
    return repo.to_response(goal)


@router.post("", response_model=GoalResponse, status_code=status.HTTP_201_CREATED)
def create_goal(
    payload: GoalCreate,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_superuser),
):
    repo = GoalRepository(db)
    if repo.get_by_code(payload.code):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Deze code bestaat al")
    return repo.to_response(repo.create(payload))


@router.put("/{goal_id}", response_model=GoalResponse)
def update_goal(
    goal_id: int,
    payload: GoalUpdate,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_superuser),
):
    repo = GoalRepository(db)
    goal = repo.update(goal_id, payload)
    if not goal:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Doel niet gevonden")
    return repo.to_response(goal)


@router.delete("/{goal_id}", response_model=GoalResponse)
def delete_goal(
    goal_id: int,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_superuser),
):
    repo = GoalRepository(db)
    goal = repo.delete(goal_id)
    if not goal:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Doel niet gevonden")
    return repo.to_response(goal)


@router.post("/import-vo", response_model=dict)
def import_vo_goals(
    file_path: str,
    db=Depends(get_db),
    current_user: UserResponse = Depends(get_current_superuser),
):
    repo = GoalRepository(db)
    goals_data = _parse_vo_excel(file_path)
    existing_codes = {g.code for g in repo.get_all(goal_type="VO")}
    new_goals = [g for g in goals_data if g["code"] not in existing_codes]
    skipped = len(goals_data) - len(new_goals)
    if new_goals:
        repo.bulk_create(new_goals)
    return {
        "imported": len(new_goals),
        "skipped": skipped,
        "total": len(goals_data),
    }
