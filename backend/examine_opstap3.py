from pathlib import Path
from openpyxl import load_workbook

base = Path("/home/davidlinux/PycharmProjects/ObsApp/AnalysisDev/op_stap")

all_codes = {}
for fname, subject in [("LP-Ned_com_B_S.xlsx", "Nederlands"), ("LP-wiskunde_B_S (1).xlsx", "Wiskunde")]:
    path = base / fname
    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[4]:
            continue
        lfmd = str(row[1]).strip() if row[1] else ""
        if lfmd != "K-":
            continue
        code = str(row[4]).strip()
        md = str(row[3]).strip() if row[3] else ""
        if code in all_codes:
            print(f"DUPLICATE CODE: {code} in {subject} and {all_codes[code]}")
        else:
            all_codes[code] = subject
    wb.close()

print(f"Total unique K- codes: {len(all_codes)}")

# Check how many K- rows have MD that matches a VO code
# VO codes are like "1.1.1", "1.1.2", etc. (without the K- prefix)
# MD is like "K-1.1.1", "K-2.1.1", etc.
from app.core.database import SessionLocal
from app.models.goal import Goal

db = SessionLocal()
vo_codes = {g.code for g in db.query(Goal).filter(Goal.goal_type == "VO").all()}
print(f"VO codes in DB: {len(vo_codes)}")

matched = 0
unmatched = 0
unmatched_mds = {}
for fname, subject in [("LP-Ned_com_B_S.xlsx", "Nederlands"), ("LP-wiskunde_B_S (1).xlsx", "Wiskunde")]:
    path = base / fname
    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[4]:
            continue
        lfmd = str(row[1]).strip() if row[1] else ""
        if lfmd != "K-":
            continue
        md = str(row[3]).strip() if row[3] else ""
        # Remove K- prefix to get VO code
        vo_code = md.removeprefix("K-") if md.startswith("K-") else md
        if vo_code in vo_codes:
            matched += 1
        else:
            unmatched += 1
            unmatched_mds[md] = unmatched_mds.get(md, 0) + 1
    wb.close()

print(f"Op Stap K- rows matched to VO: {matched}")
print(f"Op Stap K- rows NOT matched: {unmatched}")
print(f"Unmatched MDs (sample): {list(unmatched_mds.items())[:20]}")
db.close()
